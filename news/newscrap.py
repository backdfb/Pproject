import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

search = tk.Tk()

search.title("only naver news")
search.geometry("500x100")

display = tk.Entry(search, width=30)
display.pack()

def func(event):
    keyword = display.get()  # 수정된 부분
    print(keyword)  # 검색어 출력
    url = "https://search.naver.com/search.naver?where=news&sm=tab_jum&query={}".format(keyword)

    try:
        html = requests.get(url)
        html.raise_for_status()  # HTTP 오류가 발생하면 예외를 발생시킴
        soup = BeautifulSoup(html.text, "html.parser")
        myurls = soup.find_all('a', class_='news_tit')  # CSS 클래스 확인

        myurls_text = []
        for myurl in myurls:
            title = myurl.attrs['title']
            link = myurl.attrs['href']
            myurls_text.append((title, link))  # 제목과 링크를 튜플로 저장

        if myurls_text:  # 뉴스 기사가 있는 경우
            # 엑셀 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "뉴스 기사"

            # 열 제목 추가
            ws.append(['Title', 'Link'])

            # 데이터 추가 및 하이퍼링크 설정
            for title, link in myurls_text:
                ws.append([title, link])
                ws.cell(row=ws.max_row, column=2).hyperlink = link  # 링크 설정

            # 열 너비 조정
            for column in range(1, 3):
                ws.column_dimensions[get_column_letter(column)].width = 50

            # 엑셀 파일 저장
            wb.save('news_{}.xlsx'.format(keyword))
            messagebox.showinfo("완료", "뉴스 기사가 엑셀 파일에 저장되었습니다.")
        else:
            messagebox.showwarning("경고", "검색 결과가 없습니다.")

    except requests.exceptions.RequestException as e:
        messagebox.showerror("오류", f"HTTP 요청 오류: {e}")

display.bind('<Return>', func)

search.mainloop()
